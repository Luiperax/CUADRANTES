"""
Exportación del CUADRANTE DE FACTURACIÓN (formato del cliente).

A diferencia del cuadrante operativo, la hoja de facturación reorganiza el mes
por SERVICIO / orden de trabajo (OT): cada puesto es un servicio que se factura
por separado al cliente. Para cada servicio se listan los empleados que lo han
cubierto, con su hora de entrada y de salida por día, las horas por jornada
(12 h) y los totales, más el total diario del servicio y el total general.

Correspondencia puesto -> servicio (según el cuadrante real de NATURGY):

* F1 -> RECEPCIÓN 24 H            (MT 7-19 / TN 19-7)
* F2 -> GARITA 24 H              (MT 7-19 / TN 19-7)
* MO -> MÓVIL/RESPONSABLE 7 a 19 (solo mañana)
* EX -> EXPLANADA 6 a 18         (solo mañana, horario 6-18)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

from ..config.constantes import NOMBRES_MES, Puesto, TipoAusencia, Turno
from ..datos.modelos import Cuadrante, Trabajador
from ..dominio.calendario import CalendarioMes

# Servicios (OT) en el orden en que aparecen en la hoja del cliente.
# Clave interna, código de OT y nombre. La asignación de cada turno a un servicio
# NO es directa por puesto: depende del turno, el puesto y si es día laborable o
# fin de semana (ver _servicio_de), tal como en la facturación real del cliente.
_SERVICIOS = [
    ("RECEPCION", "01CE000579287511", "RECEPCIÓN 24 H"),
    ("GARITA", "01CE000579287521", "GARITA 24 H VSA"),
    ("MOVIL", "01CE000579287531", "MÓVIL/RESPONSABLE VSA DE 7 a 19"),
    ("EXPLANADA", "01CE000579287541", "EXPLANADA VSA de 6 a 18"),
]


def _servicio_de(turno: Turno, puesto: Puesto, es_finde: bool) -> str:
    """Servicio (OT) al que se factura un turno-puesto concreto.

    * EX               -> EXPLANADA (6-18).
    * F2 (MT o TN)     -> GARITA 24 H.
    * MO               -> RECEPCIÓN (cubre el día laborable de Recepción).
    * F1 de noche      -> RECEPCIÓN.
    * F1 de mañana     -> MÓVIL/RESPONSABLE si es laborable (lo hacen Luis y
                          Fernando); en fin de semana va a RECEPCIÓN.
    """
    if puesto is Puesto.EX:
        return "EXPLANADA"
    if puesto is Puesto.F2:
        return "GARITA"
    if puesto is Puesto.MO:
        return "RECEPCION"
    # Puesto F1
    if turno is Turno.NOCHE:
        return "RECEPCION"
    return "RECEPCION" if es_finde else "MOVIL"

_COMPUTO = 162.0

# --- Estilos ---------------------------------------------------------------
_FINO = Side(style="thin", color="808080")
_BORDE = Border(left=_FINO, right=_FINO, top=_FINO, bottom=_FINO)
_CENTRO = Alignment(horizontal="center", vertical="center")
_IZQ = Alignment(horizontal="left", vertical="center")
_EDIT = Protection(locked=False)

_AZUL = "9DC3E6"        # fin de semana
_PEACH = "FCE4D6"       # fila SUMA
_AMAR = "FFF2CC"        # totales
_CYAN = "00B0F0"        # vacaciones (V)
_GRIS = "D9D9D9"        # cabeceras de datos
_NEGRO = "000000"

_F_NORM = Font(name="Arial", size=8)
_F_NEG = Font(name="Arial", size=8, bold=True)
_F_PEQ = Font(name="Arial", size=7)
_F_TIT = Font(name="Arial", size=11, bold=True)
_F_OT = Font(name="Arial", size=9, bold=True, color="FFFFFF")
_F_OTCOD = Font(name="Arial", size=9, bold=True, color="FF0000")


def _relleno(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _entrada_salida(turno: Turno, puesto: Puesto) -> tuple[int, int]:
    """Hora de entrada y de salida según turno y puesto."""
    if puesto is Puesto.EX:      # Explanada 6-18 (siempre de mañana).
        return 6, 18
    if turno is Turno.NOCHE:     # Noche 19-7.
        return 19, 7
    return 7, 19                 # Mañana 7-19.


def rangos_vacaciones(cuadrante: Cuadrante, trabajadores: dict[int, Trabajador],
                      calendario: CalendarioMes) -> list[tuple[str, int, int]]:
    """Devuelve (nombre, día_inicio, día_fin) de cada periodo de vacaciones del mes."""
    res = []
    ids = cuadrante.trabajadores_ids or list(trabajadores.keys())
    for tid in ids:
        vac = [d for d in calendario.dias
               if (a := cuadrante.obtener(tid, d)) and a.ausencia is TipoAusencia.VACACIONES]
        if not vac:
            continue
        ini = prev = vac[0]
        rangos = []
        for d in vac[1:]:
            if d == prev + 1:
                prev = d
            else:
                rangos.append((ini, prev)); ini = prev = d
        rangos.append((ini, prev))
        nombre = trabajadores[tid].nombre if tid in trabajadores else str(tid)
        for a, b in rangos:
            res.append((nombre, a, b))
    return res


def construir_datos_facturacion(cuadrante: Cuadrante, trabajadores: dict[int, Trabajador],
                                calendario: CalendarioMes) -> dict:
    """Estructura lógica de la facturación (la usan el Excel y la vista en pantalla).

    Devuelve ``{'servicios': [...], 'total_general_dia': [...], 'total_general': ...}``.
    Cada servicio: ``{'codigo','nombre','empleados','totales_dia','total'}``.
    Cada empleado: ``{'nombre','celdas':[{entrada,salida,suma,vac,finde}...],'total','dif'}``.
    """
    dias = calendario.dias
    ids = cuadrante.trabajadores_ids or list(trabajadores.keys())

    def serv_dia(tid, dia):
        """Servicio de la asignación de un trabajador ese día, o None."""
        a = cuadrante.obtener(tid, dia)
        if a and a.es_trabajo:
            return _servicio_de(a.turno, a.puesto, calendario.es_fin_de_semana(dia))
        return None

    servicios = []
    for clave, codigo, nombre in _SERVICIOS:
        empleados = []
        for tid in ids:
            if not any(serv_dia(tid, d) == clave for d in dias):
                continue
            celdas, total = [], 0
            for dia in dias:
                a = cuadrante.obtener(tid, dia)
                cel = {"entrada": "", "salida": "", "suma": "", "vac": False,
                       "finde": calendario.es_fin_de_semana(dia)}
                if serv_dia(tid, dia) == clave:
                    e, s = _entrada_salida(a.turno, a.puesto)
                    cel["entrada"], cel["salida"], cel["suma"] = e, s, 12
                    total += 12
                elif a and a.ausencia is TipoAusencia.VACACIONES:
                    cel["entrada"] = cel["salida"] = "V"
                    cel["vac"] = True
                celdas.append(cel)
            nombre_t = trabajadores[tid].nombre if tid in trabajadores else str(tid)
            empleados.append({"id": tid, "nombre": nombre_t, "celdas": celdas,
                              "total": total, "dif": total - _COMPUTO})
        tot_dia = [sum(12 for e in empleados if serv_dia(e["id"], dia) == clave) for dia in dias]
        servicios.append({"codigo": codigo, "nombre": nombre, "clave": clave,
                          "empleados": empleados, "totales_dia": tot_dia, "total": sum(tot_dia)})
    tg = []
    for dia in dias:
        h = sum(12 for tid in ids if (a := cuadrante.obtener(tid, dia)) and a.es_trabajo)
        tg.append(h)
    return {"servicios": servicios, "total_general_dia": tg, "total_general": sum(tg)}


class ExportadorFacturacion:
    """Genera el cuadrante de facturación en Excel, agrupado por servicio."""

    def __init__(self, cuadrante: Cuadrante, trabajadores: dict[int, Trabajador],
                 festivos: set | None = None):
        self.cuadrante = cuadrante
        self.trabajadores = trabajadores
        self.calendario = CalendarioMes(cuadrante.anio, cuadrante.mes, festivos or set())
        self.datos = construir_datos_facturacion(cuadrante, trabajadores, self.calendario)
        self.n_dias = self.calendario.numero_dias
        self.col_dia0 = 3                       # primera columna de día (C)
        self.col_tot = self.col_dia0 + self.n_dias      # TOTALES
        self.col_dif = self.col_tot + 1                 # diferencia / etiquetas

    # ------------------------------------------------------------------
    def exportar(self, ruta: str | Path) -> Path:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Facturación"
        hoja.sheet_properties.pageSetUpPr.fitToPage = True
        hoja.page_setup.orientation = "landscape"
        hoja.page_setup.fitToWidth = 1
        hoja.page_setup.fitToHeight = 0

        fila = self._cabecera(hoja)
        servicios = self.datos["servicios"]
        for i, serv in enumerate(servicios):
            fila = self._bloque_servicio(hoja, fila, serv)
            if i < len(servicios) - 1:
                fila += 1                 # separación solo ENTRE servicios
        fila = self._total_general(hoja, fila)   # total pegado al último servicio
        self._pie(hoja, fila)
        self._anchos(hoja)

        ruta = Path(ruta)
        ruta.parent.mkdir(parents=True, exist_ok=True)
        libro.save(str(ruta))
        return ruta

    # ------------------------------------------------------------------
    def _cel(self, hoja, fila, col, valor="", fuente=_F_NORM, relleno=None,
             alin=_CENTRO, borde=True):
        c = hoja.cell(row=fila, column=col, value=valor)
        c.font = fuente
        c.alignment = alin
        c.protection = _EDIT
        if borde:
            c.border = _BORDE
        if relleno:
            c.fill = _relleno(relleno)
        return c

    def _caja(self, hoja, r1, c1, r2, c2, texto="", fuente=_F_NORM, relleno=None, alin=_CENTRO):
        """Escribe una caja combinada con borde en todo el rango."""
        if (r1, c1) != (r2, c2):
            hoja.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        self._cel(hoja, r1, c1, texto, fuente=fuente, relleno=relleno, alin=alin)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                hoja.cell(row=r, column=c).border = _BORDE

    def _cabecera(self, hoja) -> int:
        mes = NOMBRES_MES[self.cuadrante.mes].upper()
        izq = Alignment(horizontal="left", vertical="center")
        # Alturas de fila para que la cabecera respire.
        for r in (2, 3, 4):
            hoja.row_dimensions[r].height = 22

        # Fila 1: título del documento y referencia (arriba a la derecha).
        self._cel(hoja, 1, 1, "OK - REPUESTA", fuente=_F_PEQ, borde=False, alin=izq)
        self._caja(hoja, 1, self.col_tot - 3, 1, self.col_dif,
                   "Código Documento: MD-ES-SISVG-VA-03  ·  Edición: 01",
                   fuente=_F_PEQ, alin=_CENTRO)

        # Etiquetas (columna A) + valores (cajas combinadas amplias y legibles).
        et = Font(name="Arial", size=9, bold=True)
        val_g = Font(name="Arial", size=13, bold=True)   # NATURGY, MES
        val_m = Font(name="Arial", size=10, bold=True)    # centro
        # Columnas: A=etiqueta izq; B..M=valor izq; N..P=etiqueta der; Q..T=valor der.
        c_val1a, c_val1b = 2, 13
        c_lab2a, c_lab2b = 14, 17
        c_val2a, c_val2b = 18, 22

        self._caja(hoja, 2, 1, 2, 1, "CLIENTE:", fuente=et, relleno=_GRIS, alin=izq)
        self._caja(hoja, 2, c_val1a, 2, c_val1b, "NATURGY", fuente=val_g)
        self._caja(hoja, 2, c_lab2a, 2, c_lab2b, "CD. CENTRO:", fuente=et, relleno=_GRIS, alin=izq)
        self._caja(hoja, 2, c_val2a, 2, c_val2b, "", fuente=val_m)

        self._caja(hoja, 3, 1, 3, 1, "CENTRO:", fuente=et, relleno=_GRIS, alin=izq)
        self._caja(hoja, 3, c_val1a, 3, c_val1b, "EDIFICIO AVENIDA DE SAN LUIS", fuente=val_m)
        self._caja(hoja, 3, c_lab2a, 3, c_lab2b, "MES:", fuente=et, relleno=_GRIS, alin=izq)
        self._caja(hoja, 3, c_val2a, 3, c_val2b, f"{mes} {self.cuadrante.anio}", fuente=val_g)

        self._caja(hoja, 4, c_lab2a, 4, c_val2b, "SIN ARMA", fuente=et, relleno=_GRIS)
        return 6

    def _fila_dias(self, hoja, fila) -> int:
        """Escribe la fila de letras de día y la de números. Devuelve la siguiente."""
        self._cel(hoja, fila, 1, "EMPLEADO", fuente=_F_NEG, relleno=_GRIS, alin=_IZQ)
        self._cel(hoja, fila + 1, self.col_tot, "TOTALES", fuente=_F_NEG, relleno=_GRIS)
        for dia in self.calendario.dias:
            col = self.col_dia0 + dia - 1
            finde = self.calendario.es_fin_de_semana(dia)
            rel = _AZUL if finde else None
            self._cel(hoja, fila, col, self.calendario.letra_dia(dia), fuente=_F_NEG, relleno=rel)
            self._cel(hoja, fila + 1, col, dia, fuente=_F_NEG, relleno=rel)
        return fila + 2

    def _bloque_servicio(self, hoja, fila, serv) -> int:
        fila = self._fila_dias(hoja, fila)
        # Cabecera negra del servicio (código en rojo + nombre en blanco).
        self._cel(hoja, fila, 1, serv["codigo"], fuente=_F_OTCOD, relleno=_NEGRO, alin=_IZQ)
        self._cel(hoja, fila, 2, serv["nombre"], fuente=_F_OT, relleno=_NEGRO, alin=_IZQ)
        for col in range(3, self.col_dif + 1):
            self._cel(hoja, fila, col, "", relleno=_NEGRO)
        fila += 1

        for emp in serv["empleados"]:
            fila = self._filas_empleado(hoja, fila, emp)

        # Total diario del servicio.
        self._cel(hoja, fila, 1, "", borde=False)
        self._cel(hoja, fila, 2, "", borde=False)
        for i, dia in enumerate(self.calendario.dias):
            col = self.col_dia0 + dia - 1
            self._cel(hoja, fila, col, serv["totales_dia"][i] or "", fuente=_F_NEG, relleno=_AMAR)
        self._cel(hoja, fila, self.col_tot, serv["total"], fuente=_F_NEG, relleno=_AMAR)
        return fila + 1

    def _filas_empleado(self, hoja, fila, emp) -> int:
        f_ent, f_sal, f_sum = fila, fila + 1, fila + 2
        # Nombre combinado sobre las tres filas.
        hoja.merge_cells(start_row=f_ent, start_column=1, end_row=f_sum, end_column=1)
        self._cel(hoja, f_ent, 1, emp["nombre"], fuente=_F_NEG, alin=_IZQ)
        self._cel(hoja, f_ent, 2, "HORA DE ENTRADA", fuente=_F_PEQ, alin=_IZQ)
        self._cel(hoja, f_sal, 2, "HORA DE SALIDA", fuente=_F_PEQ, alin=_IZQ)
        self._cel(hoja, f_sum, 2, "SUMA", fuente=_F_PEQ, alin=_IZQ)

        for i, cel in enumerate(emp["celdas"]):
            col = self.col_dia0 + i
            rel_base = _AZUL if cel["finde"] else None
            rel_ev = _CYAN if cel["vac"] else rel_base
            rel_sum = rel_base if cel["vac"] else _PEACH
            self._cel(hoja, f_ent, col, cel["entrada"], fuente=_F_NEG, relleno=rel_ev)
            self._cel(hoja, f_sal, col, cel["salida"], fuente=_F_NEG, relleno=rel_ev)
            self._cel(hoja, f_sum, col, cel["suma"], fuente=_F_PEQ, relleno=rel_sum)
        # Etiquetas y totales a la derecha.
        self._cel(hoja, f_ent, self.col_dif, "HORAS", fuente=_F_PEQ, borde=False)
        self._cel(hoja, f_sal, self.col_dif, "EXTRAS", fuente=_F_PEQ, borde=False)
        self._cel(hoja, f_sum, self.col_tot, emp["total"] or "", fuente=_F_NEG, relleno=_AMAR)
        self._cel(hoja, f_sum, self.col_dif, f"{emp['dif']:.2f}".replace(".", ","),
                  fuente=_F_NEG, borde=False)
        return f_sum + 1

    def _total_general(self, hoja, fila) -> int:
        self._cel(hoja, fila, 1, "TOTAL GENERAL", fuente=_F_NEG, relleno=_AMAR, alin=_IZQ)
        self._cel(hoja, fila, 2, "", relleno=_AMAR)
        for i, dia in enumerate(self.calendario.dias):
            col = self.col_dia0 + dia - 1
            self._cel(hoja, fila, col, self.datos["total_general_dia"][i],
                      fuente=_F_NEG, relleno=_AMAR)
        self._cel(hoja, fila, self.col_tot, self.datos["total_general"],
                  fuente=_F_NEG, relleno=_AMAR)
        self._cel(hoja, fila, self.col_dif, "", borde=False)
        return fila + 1

    def _pie(self, hoja, fila) -> None:
        """Leyenda de códigos de servicio (OT) y notas de vacaciones, como el original."""
        mes = NOMBRES_MES[self.cuadrante.mes]
        fila += 2
        self._cel(hoja, fila, 1, "DETALLE DE LOS SERVICIOS (OT)", fuente=_F_NEG,
                  borde=False, alin=_IZQ)
        fila += 1
        rojo = Font(name="Arial", size=8, bold=True, color="FF0000")
        for puesto, codigo, nombre in _SERVICIOS:
            self._cel(hoja, fila, 1, codigo, fuente=rojo, borde=False, alin=_IZQ)
            hoja.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=14)
            self._cel(hoja, fila, 2, f"esta OT corresponde a {nombre}",
                      fuente=_F_NORM, borde=False, alin=_IZQ)
            fila += 1
        vacs = rangos_vacaciones(self.cuadrante, self.trabajadores, self.calendario)
        if vacs:
            fila += 1
            for nombre, ini, fin in vacs:
                hoja.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=16)
                texto = (f"La VS {nombre} disfruta de vacaciones del {ini} al {fin} "
                         f"de {mes}, ambos inclusive.")
                self._cel(hoja, fila, 1, texto, fuente=_F_NORM, borde=False, alin=_IZQ)
                fila += 1

    def _anchos(self, hoja) -> None:
        hoja.column_dimensions["A"].width = 28
        hoja.column_dimensions["B"].width = 14
        for dia in self.calendario.dias:
            hoja.column_dimensions[get_column_letter(self.col_dia0 + dia - 1)].width = 3.5
        hoja.column_dimensions[get_column_letter(self.col_tot)].width = 8
        hoja.column_dimensions[get_column_letter(self.col_dif)].width = 9


class ExportadorFacturacionPDF:
    """Genera el cuadrante de facturación en PDF (A3 apaisado), mismo contenido."""

    def __init__(self, cuadrante: Cuadrante, trabajadores: dict[int, Trabajador],
                 festivos: set | None = None):
        self.cuadrante = cuadrante
        self.trabajadores = trabajadores
        self.calendario = CalendarioMes(cuadrante.anio, cuadrante.mes, festivos or set())

    def exportar(self, ruta: str | Path) -> Path:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet

        cal = self.calendario
        dias = cal.dias
        nd = len(dias)
        datos = construir_datos_facturacion(self.cuadrante, self.trabajadores, cal)

        azul = colors.HexColor("#9DC3E6"); peach = colors.HexColor("#FCE4D6")
        amar = colors.HexColor("#FFF2CC"); cyan = colors.HexColor("#00B0F0")
        gris = colors.HexColor("#D9D9D9")

        filas, estilos = [], []
        estilos += [("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 4.5),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]
        col_d0 = 2
        col_tot = col_d0 + nd
        col_dif = col_tot + 1

        def sombrear_finde(ri):
            for i, dia in enumerate(dias):
                if cal.es_fin_de_semana(dia):
                    estilos.append(("BACKGROUND", (col_d0 + i, ri), (col_d0 + i, ri), azul))

        r = 0
        for s in datos["servicios"]:
            # Cabecera de días (letras + números).
            fila_l = [""] * (col_dif + 1); fila_n = [""] * (col_dif + 1)
            fila_l[0] = "EMPLEADO"
            for i, dia in enumerate(dias):
                fila_l[col_d0 + i] = cal.letra_dia(dia); fila_n[col_d0 + i] = str(dia)
            fila_n[col_tot] = "TOTALES"
            filas.append(fila_l); sombrear_finde(r)
            estilos.append(("BACKGROUND", (0, r), (0, r), gris)); estilos.append(("FONTNAME", (0, r), (-1, r + 1), "Helvetica-Bold")); r += 1
            filas.append(fila_n); sombrear_finde(r)
            estilos.append(("BACKGROUND", (col_tot, r), (col_tot, r), gris)); r += 1
            # Cabecera negra del servicio.
            filaot = [""] * (col_dif + 1); filaot[0] = s["codigo"]; filaot[1] = s["nombre"]
            filas.append(filaot)
            estilos += [("BACKGROUND", (0, r), (-1, r), colors.black),
                        ("TEXTCOLOR", (0, r), (0, r), colors.red),
                        ("TEXTCOLOR", (1, r), (1, r), colors.white),
                        ("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"),
                        ("SPAN", (1, r), (col_dif, r)), ("ALIGN", (0, r), (1, r), "LEFT")]
            r += 1
            for emp in s["empleados"]:
                fe = [""] * (col_dif + 1); fs = [""] * (col_dif + 1); fu = [""] * (col_dif + 1)
                fe[0] = emp["nombre"]; fe[1] = "H. ENTRADA"; fs[1] = "H. SALIDA"; fu[1] = "SUMA"
                fe[col_dif] = "HORAS"; fs[col_dif] = "EXTRAS"
                for i, cel in enumerate(emp["celdas"]):
                    fe[col_d0 + i] = cel["entrada"]; fs[col_d0 + i] = cel["salida"]
                    fu[col_d0 + i] = cel["suma"]
                    if cel["vac"]:
                        estilos.append(("BACKGROUND", (col_d0 + i, r), (col_d0 + i, r + 1), cyan))
                    elif cel["suma"]:
                        estilos.append(("BACKGROUND", (col_d0 + i, r + 2), (col_d0 + i, r + 2), peach))
                fu[col_tot] = emp["total"] or ""
                fu[col_dif] = f"{emp['dif']:.2f}".replace(".", ",")
                filas += [fe, fs, fu]
                sombrear_finde(r); sombrear_finde(r + 1)
                estilos += [("SPAN", (0, r), (0, r + 2)), ("ALIGN", (0, r), (1, r + 2), "LEFT"),
                            ("FONTNAME", (0, r), (0, r), "Helvetica-Bold"),
                            ("BACKGROUND", (col_tot, r + 2), (col_tot, r + 2), amar),
                            ("FONTNAME", (col_tot, r + 2), (col_dif, r + 2), "Helvetica-Bold")]
                r += 3
            # Total diario del servicio.
            ftot = [""] * (col_dif + 1)
            for i in range(nd):
                ftot[col_d0 + i] = s["totales_dia"][i] or ""
            ftot[col_tot] = s["total"]
            filas.append(ftot)
            estilos += [("BACKGROUND", (col_d0, r), (col_tot, r), amar),
                        ("FONTNAME", (0, r), (-1, r), "Helvetica-Bold")]
            r += 1

        # Total general.
        fg = [""] * (col_dif + 1); fg[0] = "TOTAL GENERAL"
        for i in range(nd):
            fg[col_d0 + i] = datos["total_general_dia"][i]
        fg[col_tot] = datos["total_general"]
        filas.append(fg)
        estilos += [("BACKGROUND", (0, r), (col_tot, r), amar),
                    ("FONTNAME", (0, r), (-1, r), "Helvetica-Bold"), ("ALIGN", (0, r), (0, r), "LEFT")]

        anchos = [92, 40] + [12] * nd + [24, 28]
        tabla = Table(filas, colWidths=anchos, repeatRows=0)
        tabla.setStyle(TableStyle(estilos))

        ruta = Path(ruta); ruta.parent.mkdir(parents=True, exist_ok=True)
        mes = NOMBRES_MES[self.cuadrante.mes].upper()

        # Cabecera en cajas (CLIENTE / CENTRO / MES), legible y como el original.
        hfilas = [
            ["CLIENTE:", "NATURGY", "CD. CENTRO:", "", "Código Documento:\nMD-ES-SISVG-VA-03 · Ed. 01"],
            ["CENTRO:", "EDIFICIO AVENIDA DE SAN LUIS", "MES:", f"{mes} {self.cuadrante.anio}", ""],
            ["", "", "SIN ARMA", "", ""],
        ]
        htabla = Table(hfilas, colWidths=[80, 250, 80, 130, 170], rowHeights=[20, 20, 18])
        htabla.setStyle(TableStyle([
            ("BOX", (0, 0), (3, 2), 0.6, colors.black),
            ("GRID", (0, 0), (3, 2), 0.4, colors.grey),
            ("BACKGROUND", (0, 0), (0, 1), gris),
            ("BACKGROUND", (2, 0), (2, 2), gris),
            ("SPAN", (2, 2), (3, 2)),        # SIN ARMA
            ("SPAN", (4, 0), (4, 2)),        # referencia
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTSIZE", (1, 0), (1, 0), 14),   # NATURGY
            ("FONTSIZE", (3, 1), (3, 1), 14),   # MES valor
            ("FONTSIZE", (1, 1), (1, 1), 10),   # centro
            ("FONTSIZE", (4, 0), (4, 0), 6.5), ("FONTNAME", (4, 0), (4, 0), "Helvetica"),
            ("ALIGN", (1, 0), (1, 1), "CENTER"), ("ALIGN", (3, 1), (3, 1), "CENTER"),
            ("ALIGN", (2, 2), (3, 2), "CENTER"), ("ALIGN", (4, 0), (4, 0), "CENTER"),
            ("ALIGN", (0, 0), (0, 1), "LEFT"), ("ALIGN", (2, 0), (2, 1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (4, 0), (4, 2), 0, colors.white),
            ("LINEAFTER", (3, 0), (3, 2), 0, colors.white),
        ]))
        doc = SimpleDocTemplate(str(ruta), pagesize=landscape(A3),
                                leftMargin=8 * mm, rightMargin=8 * mm,
                                topMargin=8 * mm, bottomMargin=8 * mm)

        from reportlab.lib.styles import ParagraphStyle
        p_txt = ParagraphStyle("pie", fontName="Helvetica", fontSize=7.5, leading=10)
        p_tit = ParagraphStyle("pietit", fontName="Helvetica-Bold", fontSize=8, leading=11)
        elementos = [htabla, Spacer(1, 4 * mm), tabla, Spacer(1, 5 * mm),
                     Paragraph("DETALLE DE LOS SERVICIOS (OT)", p_tit)]
        for _p, codigo, nombre in _SERVICIOS:
            elementos.append(Paragraph(f"<b>{codigo}</b> — esta OT corresponde a {nombre}", p_txt))
        vacs = rangos_vacaciones(self.cuadrante, self.trabajadores, cal)
        if vacs:
            elementos.append(Spacer(1, 3 * mm))
            mes_min = NOMBRES_MES[self.cuadrante.mes]
            for nombre, ini, fin in vacs:
                elementos.append(Paragraph(
                    f"La VS {nombre} disfruta de vacaciones del {ini} al {fin} "
                    f"de {mes_min}, ambos inclusive.", p_txt))
        doc.build(elementos)
        return ruta
