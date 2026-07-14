"""
Exportación del cuadrante a Excel replicando fielmente el formato NATURGY.

El objetivo es que el fichero generado sea prácticamente indistinguible del
cuadrante original: mismos colores, bordes, celdas combinadas, leyenda de turnos,
cabecera, columnas H.T./H.E./H.N., resaltado de fines de semana y disposición de
dos filas por trabajador (turno arriba, puesto abajo).

Se emplea OpenPyXL por ser la biblioteca de referencia para leer y escribir
ficheros .xlsx con control total sobre estilos, algo imprescindible para lograr
una reproducción visual exacta.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from ..config.constantes import (
    HORARIOS,
    LEYENDA_CABECERA,
    NOMBRES_MES,
    Colores,
    Puesto,
    TipoAusencia,
    Turno,
)
from ..datos.modelos import Cuadrante, Trabajador
from ..dominio.calendario import CalendarioMes
from ..dominio.computos import calcular_resumenes


# --- Estilos reutilizables -------------------------------------------------
_LADO_FINO = Side(style="thin", color=Colores.BORDE)
_BORDE = Border(left=_LADO_FINO, right=_LADO_FINO, top=_LADO_FINO, bottom=_LADO_FINO)
_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)
_IZQUIERDA = Alignment(horizontal="left", vertical="center")

# Todas las celdas se marcan como NO bloqueadas para que el fichero se declare
# explícitamente editable. Por defecto openpyxl marca las celdas como «locked»,
# lo que, aunque solo tiene efecto si se protege la hoja (aquí nunca se protege),
# algunos visores (sobre todo de móvil) interpretan como «solo lectura». Así se
# garantiza que el cuadrante se pueda modificar a mano en cualquier programa.
_EDITABLE = Protection(locked=False)

_FUENTE_TITULO = Font(name="Arial", size=14, bold=True)
_FUENTE_NORMAL = Font(name="Arial", size=8)
_FUENTE_NEGRITA = Font(name="Arial", size=8, bold=True)
_FUENTE_CELDA = Font(name="Arial", size=7, bold=True)


def _relleno(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


class ExportadorExcel:
    """Genera el fichero .xlsx del cuadrante con el formato NATURGY."""

    def __init__(self, cuadrante: Cuadrante, trabajadores: dict[int, Trabajador]):
        self.cuadrante = cuadrante
        self.trabajadores = trabajadores
        self.calendario = CalendarioMes(cuadrante.anio, cuadrante.mes)
        self.resumenes = calcular_resumenes(cuadrante, trabajadores, self.calendario)
        # Columna inicial de los días (tras la columna de nombres).
        self.col_nombre = 1
        self.col_dia_inicial = 2
        self.n_dias = self.calendario.numero_dias
        self.col_ht = self.col_dia_inicial + self.n_dias
        self.col_he = self.col_ht + 1
        self.col_hn = self.col_he + 1

    # ------------------------------------------------------------------
    def exportar(self, ruta: str | Path) -> Path:
        libro = Workbook()
        hoja = libro.active
        hoja.title = NOMBRES_MES[self.cuadrante.mes].capitalize()

        self._configurar_pagina(hoja)
        fila = self._escribir_cabecera(hoja)
        fila = self._escribir_leyenda(hoja, fila)
        fila = self._escribir_filas_encabezado(hoja, fila)
        fila = self._escribir_trabajadores(hoja, fila)
        self._escribir_computo_diario(hoja, fila)
        self._ajustar_anchos(hoja)

        ruta = Path(ruta)
        ruta.parent.mkdir(parents=True, exist_ok=True)
        libro.save(str(ruta))
        return ruta

    # ------------------------------------------------------------------
    def _configurar_pagina(self, hoja) -> None:
        # La hoja NUNCA se protege: el cuadrante debe poder editarse a mano.
        hoja.protection.sheet = False
        hoja.page_setup.orientation = "landscape"
        hoja.page_setup.fitToWidth = 1
        hoja.page_setup.fitToHeight = 0
        hoja.sheet_properties.pageSetUpPr.fitToPage = True
        hoja.page_margins = PageMargins(left=0.2, right=0.2, top=0.3, bottom=0.3)
        hoja.print_options.horizontalCentered = True

    def _celda(self, hoja, fila, col, valor="", fuente=_FUENTE_NORMAL,
               relleno=None, alineacion=_CENTRO, borde=True):
        celda = hoja.cell(row=fila, column=col, value=valor)
        celda.font = fuente
        celda.alignment = alineacion
        celda.protection = _EDITABLE  # Celda editable en cualquier programa.
        if borde:
            celda.border = _BORDE
        if relleno:
            celda.fill = _relleno(relleno)
        return celda

    def _escribir_cabecera(self, hoja) -> int:
        # Título de empresa y sede.
        self._celda(hoja, 1, self.col_nombre, self.cuadrante.empresa,
                    fuente=_FUENTE_TITULO, alineacion=_CENTRO, borde=False)
        self._celda(hoja, 2, self.col_nombre, self.cuadrante.sede,
                    fuente=_FUENTE_NEGRITA, alineacion=_CENTRO, borde=False)
        # Se formatea solo el número con coma decimal (estilo español), sin tocar «C.M.».
        cm_texto = f"{self.cuadrante.computo_mensual:.2f}".replace(".", ",")
        self._celda(hoja, 3, self.col_nombre, f"C.M.= {cm_texto}",
                    fuente=_FUENTE_NORMAL, alineacion=_IZQUIERDA, borde=False)
        # Leyenda descriptiva combinada.
        hoja.merge_cells(start_row=1, start_column=3, end_row=1, end_column=min(14, self.col_ht - 1))
        self._celda(hoja, 1, 3, LEYENDA_CABECERA, fuente=_FUENTE_NEGRITA, borde=False)
        # Fecha de actualización a la derecha.
        self._celda(hoja, 3, self.col_hn, f"Actualizado {date.today().strftime('%d-%b-%y')}",
                    fuente=_FUENTE_NORMAL, borde=False)
        return 1

    def _escribir_leyenda(self, hoja, fila_base: int) -> int:
        """Escribe la tabla de leyenda de turnos en la esquina superior derecha."""
        # Pares (turno, puesto) y su horario, tal y como en el original.
        entradas = [
            (Turno.MANANA_TARDE, Puesto.F1), (Turno.NOCHE, Puesto.F1), (Turno.MANANA_TARDE, Puesto.MO),
            (Turno.MANANA_TARDE, Puesto.F2), (Turno.NOCHE, Puesto.F2), (Turno.MANANA_TARDE, Puesto.EX),
        ]
        col_ini = max(self.col_hn - 8, self.col_dia_inicial)
        for i, (turno, puesto) in enumerate(entradas):
            fila = 1 + (i % 2)
            desplaz = (i // 2) * 3
            etiqueta = f"{turno.value} {puesto.value}"
            horario = str(HORARIOS[(turno, puesto)])
            relleno = Colores.TURNO_NOCHE if turno.es_nocturno else Colores.BLANCO
            self._celda(hoja, fila, col_ini + desplaz, etiqueta, fuente=_FUENTE_CELDA, relleno=relleno)
            hoja.merge_cells(start_row=fila, start_column=col_ini + desplaz + 1,
                             end_row=fila, end_column=col_ini + desplaz + 2)
            self._celda(hoja, fila, col_ini + desplaz + 1, horario, fuente=_FUENTE_NORMAL)
        return 4  # Deja una fila en blanco antes de la rejilla.

    def _escribir_filas_encabezado(self, hoja, fila_base: int) -> int:
        """Escribe la fila del nombre del mes + números de día y la de letras."""
        fila_num = fila_base + 1
        fila_letra = fila_num + 1

        # Celda del mes (verde) combinada sobre las dos filas de encabezado.
        hoja.merge_cells(start_row=fila_num, start_column=self.col_nombre,
                         end_row=fila_letra, end_column=self.col_nombre)
        self._celda(hoja, fila_num, self.col_nombre,
                    f"{NOMBRES_MES[self.cuadrante.mes]} {self.cuadrante.anio}",
                    fuente=_FUENTE_NEGRITA, relleno=Colores.CABECERA_MES)

        for dia in self.calendario.dias:
            col = self.col_dia_inicial + dia - 1
            finde = self.calendario.es_festivo_o_finde(dia)
            relleno = Colores.FIN_DE_SEMANA if finde else Colores.BLANCO
            self._celda(hoja, fila_num, col, dia, fuente=_FUENTE_CELDA, relleno=relleno)
            self._celda(hoja, fila_letra, col, self.calendario.letra_dia(dia),
                        fuente=_FUENTE_CELDA, relleno=relleno)

        # Cabeceras H.T./H.E./H.N. combinadas sobre las dos filas.
        for col, texto in ((self.col_ht, "H.T."), (self.col_he, "H.E."), (self.col_hn, "H.N")):
            hoja.merge_cells(start_row=fila_num, start_column=col, end_row=fila_letra, end_column=col)
            self._celda(hoja, fila_num, col, texto, fuente=_FUENTE_CELDA)
        return fila_letra + 1

    def _color_ausencia(self, ausencia: TipoAusencia | None) -> str | None:
        if ausencia is TipoAusencia.VACACIONES:
            return Colores.VACACIONES
        if ausencia is TipoAusencia.BAJA_MEDICA:
            return Colores.BAJA
        return None

    def _escribir_trabajadores(self, hoja, fila_base: int) -> int:
        fila = fila_base
        ids = self.cuadrante.trabajadores_ids or list(self.trabajadores.keys())
        for trabajador_id in ids:
            trabajador = self.trabajadores.get(trabajador_id)
            nombre = trabajador.nombre if trabajador else f"Trabajador {trabajador_id}"
            fila_turno = fila
            fila_puesto = fila + 1

            # Nombre combinado sobre las dos filas.
            hoja.merge_cells(start_row=fila_turno, start_column=self.col_nombre,
                             end_row=fila_puesto, end_column=self.col_nombre)
            self._celda(hoja, fila_turno, self.col_nombre, nombre,
                        fuente=_FUENTE_NEGRITA, alineacion=_IZQUIERDA)

            for dia in self.calendario.dias:
                col = self.col_dia_inicial + dia - 1
                asignacion = self.cuadrante.obtener(trabajador_id, dia)
                finde = self.calendario.es_festivo_o_finde(dia)
                relleno_base = Colores.FIN_DE_SEMANA if finde else Colores.BLANCO

                cod_turno = asignacion.codigo_turno() if asignacion else ""
                cod_puesto = asignacion.codigo_puesto() if asignacion else ""

                # Color de la celda de turno.
                relleno_turno = relleno_base
                if asignacion and asignacion.es_cambio_manual:
                    relleno_turno = Colores.CAMBIO
                elif asignacion and asignacion.es_noche:
                    relleno_turno = Colores.TURNO_NOCHE
                elif asignacion and asignacion.ausencia:
                    color = self._color_ausencia(asignacion.ausencia)
                    if color:
                        relleno_turno = color

                self._celda(hoja, fila_turno, col, cod_turno, fuente=_FUENTE_CELDA, relleno=relleno_turno)
                relleno_puesto = Colores.PUESTO if cod_puesto else relleno_base
                self._celda(hoja, fila_puesto, col, cod_puesto, fuente=_FUENTE_CELDA, relleno=relleno_puesto)

            # Columnas de cómputo (combinadas sobre las dos filas).
            resumen = self.resumenes.get(trabajador_id)
            ht = resumen.horas_trabajadas if resumen else 0
            he = resumen.horas_extra if resumen else 0
            hn = resumen.horas_nocturnas if resumen else 0
            for col, valor in ((self.col_ht, ht), (self.col_he, he), (self.col_hn, hn)):
                hoja.merge_cells(start_row=fila_turno, start_column=col, end_row=fila_puesto, end_column=col)
                texto = str(valor).replace(".", ",") if isinstance(valor, float) else valor
                self._celda(hoja, fila_turno, col, texto, fuente=_FUENTE_NEGRITA)

            fila += 2
        return fila

    def _escribir_computo_diario(self, hoja, fila: int) -> None:
        """Escribe la fila «COMPUTO HORAS DIARIAS» con el total de horas por día."""
        self._celda(hoja, fila, self.col_nombre, "COMPUTO HORAS DIARIAS",
                    fuente=_FUENTE_NEGRITA, alineacion=_IZQUIERDA)
        total_mes = 0
        for dia in self.calendario.dias:
            col = self.col_dia_inicial + dia - 1
            horas = 0
            for trabajador_id in (self.cuadrante.trabajadores_ids or self.trabajadores.keys()):
                asignacion = self.cuadrante.obtener(trabajador_id, dia)
                if asignacion and asignacion.es_trabajo:
                    horas += 12
            total_mes += horas
            finde = self.calendario.es_festivo_o_finde(dia)
            relleno = Colores.FIN_DE_SEMANA if finde else Colores.BLANCO
            self._celda(hoja, fila, col, horas, fuente=_FUENTE_CELDA, relleno=relleno)
        # Total mensual bajo las columnas de cómputo.
        hoja.merge_cells(start_row=fila, start_column=self.col_ht, end_row=fila, end_column=self.col_hn)
        self._celda(hoja, fila, self.col_ht, total_mes, fuente=_FUENTE_NEGRITA)

    def _ajustar_anchos(self, hoja) -> None:
        hoja.column_dimensions[get_column_letter(self.col_nombre)].width = 30
        for dia in self.calendario.dias:
            hoja.column_dimensions[get_column_letter(self.col_dia_inicial + dia - 1)].width = 3.2
        for col in (self.col_ht, self.col_he, self.col_hn):
            hoja.column_dimensions[get_column_letter(col)].width = 6
        # Altura de filas ligeramente reducida para asemejar el original.
        for fila in range(1, hoja.max_row + 1):
            hoja.row_dimensions[fila].height = 13
